'''
处理存储在Excel里面的测试用例
读取用例，写入测试结果
'''
import pprint

import openpyxl
from configs.get_path import config_file_path
from tools.do_config import DoConfig


def get_case(filepath):
    workbook = openpyxl.load_workbook(filepath)
    # 读取配置文件，根据配置文件 读取要执行的用例
    mode = eval(DoConfig.get_conf_data(config_file_path,'MODE','mode'))
    print(mode)
    testlist = []
    for key in mode:
        worksheet = workbook[key]
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        if mode[key] == 'all':
            for x in range(2, max_row+1):
                testdict1 = {'sheet': key}
                for y in range(1, max_col+1):
                    testdict1[worksheet.cell(1, y).value] = worksheet.cell(x, y).value
                testlist.append(testdict1)
        else:
            for x in mode[key]:
                testdict2 = {'sheet': key}
                for y in range(1, max_col + 1):
                    testdict2[worksheet.cell(1, y).value] = worksheet.cell(x+1, y).value
                testlist.append(testdict2)

    return testlist


def write_result(filepath, sheet, row, response, testresult):
    workbook = openpyxl.load_workbook(filepath)
    worksheet = workbook[sheet]
    worksheet[f'I{row}'] = response  # 写入接口返回
    worksheet[f'J{row}'] = testresult  # 写入测试结果，pass 或者 failed
    workbook.save(filepath)




if __name__ == '__main__':
    # write_result('/Users/meidi/PycharmProjects/AUTO_API_01/data/testdata.xlsx','Sheet1',5,'FFFF','FGGGG')

    # res = get_case('/Users/meidi/PycharmProjects/AUTO_API_01/data/testdata.xlsx')
    # pprint.pprint(res)

    from get_data import GetData
    a = getattr(GetData, 'admin_tel')
    print(a)
    setattr(GetData, 'admin_tel','123456789')
    b = getattr(GetData, 'admin_tel')
    print(b)